"""주간배차표 렌더러 — 기존 게시 양식 재현 (스펙 3.1).

현장이 수년간 학습한 레이아웃을 그대로 따른다 ("더 예쁜 표" 금지 — 스펙 7):
- 일요일 시작 주간 패널을 가로로 나란히 배치
- 각 패널: 차량번호 열 + 7일 × (순번|오전|오후)
- 출발지그룹별 행 블록 + 사이 빈 행
- 하단 고지문 2줄
"""
from __future__ import annotations

import datetime as dt

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import MonthlyRoster

WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]

NOTICE_1 = "고정 및 S/P 사원은 항상 배차변경 될 수 있으며 꼭 매일 확인바랍니다."
NOTICE_2 = (
    "고정기사님은 본인차량이 운행하지 않는 날 타차량에 배차될 수 있으니 "
    "꼼꼼히 확인하시기 바랍니다."
)

_thin = Side(style="thin")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_center = Alignment(horizontal="center", vertical="center")


def _weeks_covering_month(year: int, month: int) -> list[list[dt.date]]:
    """월 전체를 덮는 일~토 주간들."""
    import calendar as _cal

    first = dt.date(year, month, 1)
    last = dt.date(year, month, _cal.monthrange(year, month)[1])
    # 일요일 시작 (weekday(): 월=0 … 일=6)
    start = first - dt.timedelta(days=(first.weekday() + 1) % 7)
    weeks = []
    d = start
    while d <= last:
        weeks.append([d + dt.timedelta(days=i) for i in range(7)])
        d += dt.timedelta(days=7)
    return weeks


def render_weekly_xlsx(
    roster: MonthlyRoster,
    path: str,
    title: str | None = None,
    posted_on: dt.date | None = None,
) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{roster.year}년 {roster.month}월"
    title = title or f"{roster.division}노선 배차표"
    posted = posted_on or dt.date.today()

    weeks = _weeks_covering_month(roster.year, roster.month)
    panel_w = 1 + 7 * 3  # 차량 열 + 7일×3

    # 제목행
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
    ws.cell(row=1, column=panel_w * len(weeks) - 3,
            value=f"게시일 {posted.isoformat()}")

    r_date, r_head, r_body = 4, 5, 6
    red_fill = PatternFill("solid", fgColor="FFCCCC")

    for wi, week in enumerate(weeks):
        base = 1 + wi * (panel_w + 1)
        # 차량 열 헤더
        ws.cell(row=r_head, column=base, value="차량번호").alignment = _center
        for di, d in enumerate(week):
            c0 = base + 1 + di * 3
            dc = ws.cell(
                row=r_date, column=c0,
                value=f"{d.isoformat()}({WEEKDAY_KO[d.weekday()]})",
            )
            dc.alignment = _center
            dc.font = Font(bold=True)
            ws.merge_cells(
                start_row=r_date, start_column=c0,
                end_row=r_date, end_column=c0 + 2,
            )
            for j, h in enumerate(("순번", "오전", "오후")):
                hc = ws.cell(row=r_head, column=c0 + j, value=h)
                hc.alignment = _center
                hc.border = _border

        row = r_body
        for gi, g in enumerate(roster.groups):
            for v in g.vehicles:
                vc = ws.cell(row=row, column=base, value=v)
                vc.alignment = _center
                vc.border = _border
                for di, d in enumerate(week):
                    c0 = base + 1 + di * 3
                    e = roster.entries.get((d, v))
                    vals = ("", "", "")
                    if e is not None and (
                        d.month == roster.month or e.slot_label or e.am.raw
                    ):
                        vals = (
                            e.slot_label or ("" if e.slot_index is None else str(e.slot_index)),
                            e.am.driver or e.am.raw,
                            e.pm.driver or e.pm.raw,
                        )
                        if e.slot_label is None and not e.am.raw and not e.pm.raw:
                            vals = ("○", "", "")  # 휴차 (COMPACT 표기)
                    for j, val in enumerate(vals):
                        cell = ws.cell(row=row, column=c0 + j, value=val)
                        cell.alignment = _center
                        cell.border = _border
                row += 1
            if gi < len(roster.groups) - 1:
                row += 1  # 그룹 구분 빈 행

    notice_row = r_body + sum(len(g.vehicles) for g in roster.groups) \
        + len(roster.groups)
    ws.cell(row=notice_row + 1, column=1, value=NOTICE_1).font = Font(
        bold=True, color="FF0000"
    )
    ws.cell(row=notice_row + 2, column=1, value=NOTICE_2).font = Font(
        bold=True, color="FF0000"
    )

    # 열 폭
    from openpyxl.utils import get_column_letter

    for wi in range(len(weeks)):
        base = 1 + wi * (panel_w + 1)
        ws.column_dimensions[get_column_letter(base)].width = 8
        for di in range(7):
            c0 = base + 1 + di * 3
            ws.column_dimensions[get_column_letter(c0)].width = 5
            ws.column_dimensions[get_column_letter(c0 + 1)].width = 8
            ws.column_dimensions[get_column_letter(c0 + 2)].width = 8

    wb.save(path)
    return path
