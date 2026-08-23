"""주간배차표(게시용) 파서.

실존 포맷 대응 (스펙 6):
- 날짜 앵커: "YYYY-MM-DD(요일)" 문자열(지선), datetime 셀,
  "일 26 일" / "금 5/1 일" 텍스트(간선) 혼재.
- 7일 패널 × 좌우 다중 패널, 노선별 세로 블록(간선).
- 날짜 셀 기준 상대좌표로 (순번, 오전, 오후) 추출.
- 차량번호 3~4자리 숫자 행 앵커, 그룹 구분 빈 행,
  "휴"/"O휴"/"○"/공란/0 상태 구분.
"""
from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from typing import Any, Optional

import openpyxl

from ..models import CellState, DayEntry, DepotGroup, LeaveType, MonthlyRoster

ISO_DATE_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})")
# "일 26 일", "화 02 일", "금 5/1 일", "월 6/1 일"
KR_DAY_RE = re.compile(
    r"^\s*[일월화수목금토]\s+(?:(\d{1,2})\s*/\s*)?(\d{1,2})\s*일?\s*$"
)
SHEET_YM_RE = re.compile(r"(\d{4})\s*년\s*(\d{1,2})\s*월")
VEHICLE_RE = re.compile(r"^\d{3,4}$")
# 순번 라벨: "3", "가좌2", "일신5"
SLOT_RE = re.compile(r"^([가-힣]*)\s*(\d{1,2})$")
REST_MARKS = {"○", "O", "o", "0", "◯", "휴차"}


@dataclass
class DateAnchor:
    row: int
    col: int
    date: dt.date


def _parse_anchor_value(v: Any) -> Optional[dt.date]:
    """완전한 날짜를 담은 앵커(datetime 셀, ISO 문자열)만 즉시 해석."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        m = ISO_DATE_RE.match(v.strip())
        if m:
            return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _parse_kr_day(v: Any) -> Optional[tuple[Optional[int], int]]:
    """간선식 텍스트 앵커 → (월 or None, 일)."""
    if not isinstance(v, str):
        return None
    m = KR_DAY_RE.match(v)
    if not m:
        return None
    month = int(m.group(1)) if m.group(1) else None
    return month, int(m.group(2))


def find_anchors(ws, year: int, month: int) -> dict[int, list[DateAnchor]]:
    """행별 날짜 앵커 스캔. 텍스트 앵커는 시퀀스 추론으로 절대 날짜 확정.

    텍스트 앵커 규칙: 같은 행에서 열 순서대로 날짜가 하루씩 증가한다.
    - "M/D" 형이 나오면 그 달로 전환(월 경계의 명시 표기).
    - 일(day)이 직전보다 감소하면 월이 넘어간 것.
    - 첫 앵커의 일이 20을 넘고 시트 월 초라면 전월 스필오버.
    """
    raw: dict[int, list[tuple[int, Any]]] = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if _parse_anchor_value(c.value) is not None or _parse_kr_day(c.value):
                raw.setdefault(c.row, []).append((c.column, c.value))

    out: dict[int, list[DateAnchor]] = {}
    for r, cells in raw.items():
        cells.sort()
        anchors: list[DateAnchor] = []
        cur_y, cur_m = year, month
        prev_day: Optional[int] = None
        first = True
        for col, v in cells:
            d = _parse_anchor_value(v)
            if d is None:
                kr = _parse_kr_day(v)
                if kr is None:
                    continue
                m_explicit, day = kr
                if m_explicit is not None:
                    if m_explicit < cur_m and cur_m == 12 and m_explicit == 1:
                        cur_y += 1
                    cur_m = m_explicit
                elif first and day > 20 and month <= 12:
                    # 첫 앵커가 전월 말 스필오버 (예: 5월 시트의 "일 26 일")
                    cur_m = month - 1 if month > 1 else 12
                    cur_y = year if month > 1 else year - 1
                elif prev_day is not None and day < prev_day and m_explicit is None:
                    cur_m += 1
                    if cur_m > 12:
                        cur_m, cur_y = 1, cur_y + 1
                try:
                    d = dt.date(cur_y, cur_m, day)
                except ValueError:
                    continue
                prev_day = day
            else:
                prev_day = d.day
            first = False
            anchors.append(DateAnchor(r, col, d))
        # 유효 패널 행: 앵커 5개 이상 (제목행의 "게시일 …" 오탐 배제)
        if len(anchors) >= 5:
            out[r] = anchors
    return out


def classify_cell(v: Any) -> CellState:
    """오전/오후 셀 → 기사명 or 휴무유형."""
    if v is None:
        return CellState(raw="")
    s = str(v).strip()
    if not s:
        return CellState(raw="")
    if s in REST_MARKS:
        return CellState(raw=s)  # 휴차 마크 — 기사 아님
    if s in ("휴", "휴무"):
        return CellState(leave=LeaveType.REGULAR, raw=s)
    if s in ("O휴", "0휴", "o휴", "○휴"):
        return CellState(leave=LeaveType.PAID, raw=s)
    if "연차" in s:
        return CellState(leave=LeaveType.ANNUAL, raw=s)
    if "병가" in s:
        return CellState(leave=LeaveType.SICK, raw=s)
    if "사후" in s:
        return CellState(leave=LeaveType.POST, raw=s)
    if "교육" in s:
        return CellState(leave=LeaveType.EDUCATION, raw=s)
    if s in ("결행", "미운행", "운휴"):
        return CellState(raw=s)  # 결행 마커 — 기사 아님
    return CellState(driver=s, raw=s)


def parse_slot(v: Any) -> tuple[Optional[str], Optional[int]]:
    """순번 셀 → (라벨, 그룹 내 숫자 순번). 휴차/공란은 (None, None)."""
    if v is None:
        return None, None
    s = str(v).strip()
    if not s or s in REST_MARKS:
        return None, None
    if isinstance(v, float) and v == int(v):
        s = str(int(v))
    m = SLOT_RE.match(s)
    if not m:
        return None, None
    return s, int(m.group(2))


def _vehicle_at(ws_grid: dict[tuple[int, int], Any], row: int, col: int) -> Optional[str]:
    v = ws_grid.get((row, col))
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip()
    return s if VEHICLE_RE.match(s) else None


def parse_weekly_sheet(
    ws, year: int, month: int, division: str = ""
) -> MonthlyRoster:
    """주간배차표 시트 1장 → MonthlyRoster."""
    grid: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                grid[(c.row, c.column)] = c.value

    anchor_rows = find_anchors(ws, year, month)
    roster = MonthlyRoster(year=year, month=month, division=division)
    group_map: dict[str, DepotGroup] = {}

    for anchor_row, anchors in sorted(anchor_rows.items()):
        # 패널 분할: 앵커 간 열 간격 > 3 이면 새 패널
        panels: list[list[DateAnchor]] = []
        for a in anchors:
            if panels and a.col - panels[-1][-1].col <= 3:
                panels[-1].append(a)
            else:
                panels.append([a])

        # 데이터 행 수집: 헤더행(앵커행+1) 아래에서 차량번호 앵커 행 탐색
        header_row = anchor_row + 1
        vehicle_col = panels[0][0].col - 1
        data_rows: list[int] = []
        blanks = 0
        r = header_row + 1
        max_row = ws.max_row
        while r <= max_row and blanks < 4:
            vehicle = _vehicle_at(grid, r, vehicle_col)
            cell = grid.get((r, vehicle_col))
            if vehicle:
                data_rows.append(r)
                blanks = 0
            elif isinstance(cell, str) and len(cell.strip()) > 6:
                break  # 하단 고지문 도달
            else:
                blanks += 1
            r += 1

        # 그룹 분할: 연속 행 블록 단위 (빈 행 = 그룹 경계)
        blocks: list[list[int]] = []
        for row_i in data_rows:
            if blocks and row_i - blocks[-1][-1] == 1:
                blocks[-1].append(row_i)
            else:
                blocks.append([row_i])

        # 블록 제목(간선: "37번 원창동")을 앵커행 위에서 탐색
        block_title = ""
        for up in range(1, 6):
            t = grid.get((anchor_row - up, vehicle_col))
            if isinstance(t, str) and ("번" in t or "출발" in t):
                block_title = t.strip()
                break

        for bi, block in enumerate(blocks):
            vehicles = [_vehicle_at(grid, ri, vehicle_col) for ri in block]
            prefix = ""
            # 순번 라벨 접두로 그룹명 추론 (가좌/일신…)
            for ri in block:
                for panel in panels:
                    lab, _ = parse_slot(grid.get((ri, panel[0].col)))
                    if lab:
                        m = SLOT_RE.match(lab)
                        if m and m.group(1):
                            prefix = m.group(1)
                        break
                if prefix:
                    break
            gname = prefix or (block_title if block_title else f"블록{anchor_row}") + (
                f"-{bi}" if not prefix and len(blocks) > 1 and not block_title else ""
            )
            if not prefix and block_title and len(blocks) > 1:
                gname = f"{block_title}-{bi}"
            key = f"{anchor_row}:{gname}"
            if key not in group_map:
                group_map[key] = DepotGroup(name=gname, slot_prefix=prefix)
                roster.groups.append(group_map[key])
            g = group_map[key]
            for v in vehicles:
                if v and v not in g.vehicles:
                    g.vehicles.append(v)

            for ri, vehicle in zip(block, vehicles):
                if not vehicle:
                    continue
                for panel in panels:
                    for a in panel:
                        slot_label, slot_idx = parse_slot(grid.get((ri, a.col)))
                        am = classify_cell(grid.get((ri, a.col + 1)))
                        pm = classify_cell(grid.get((ri, a.col + 2)))
                        entry = DayEntry(
                            date=a.date, vehicle=vehicle,
                            slot_label=slot_label, slot_index=slot_idx,
                            am=am, pm=pm,
                        )
                        prev = roster.entries.get((a.date, vehicle))
                        # 좌우 패널 겹침(같은 날짜가 두 패널에 나올 때): 내용 있는 쪽 우선
                        if prev is None or (
                            prev.slot_index is None
                            and not prev.am.raw and not prev.pm.raw
                        ):
                            roster.entries[(a.date, vehicle)] = entry
    return roster


def parse_workbook_month(
    path: str, sheet_name: str, division: str = ""
) -> MonthlyRoster:
    m = SHEET_YM_RE.search(sheet_name)
    if not m:
        raise ValueError(f"시트명에서 연월을 찾을 수 없음: {sheet_name}")
    year, month = int(m.group(1)), int(m.group(2))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        return parse_weekly_sheet(ws, year, month, division=division)
    finally:
        wb.close()
