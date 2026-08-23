"""월간배차 시트 파서 (담당자 작업용 월 단위 그리드).

주간배차표와 담는 정보는 같지만 축이 다르다:

    주간배차표 : 행 = 차량,               열 = 날짜마다 (순번|오전|오후)
    월간배차   : 행 = (노선|순번|차번),   열 = 날짜마다 (오전|오후)

가장 큰 차이는 **순번이 열이 아니라 행 라벨**이라는 점이다. 즉 이 시트의
순번은 그 달 내내 고정된 '차량의 기본 순서'이고, 날짜별 실제 순번은
일일배차의 `조율`(로테이션 오프셋)로 결정된다. 따라서 이 포맷에서는
로테이션 규칙을 시트에서 추론할 수 없고 설정값(rotation_step)을 써야 한다
— generate.py 가 그렇게 처리한다.

실측 레이아웃 (성민 간선 16/9/3-2번, 2026):
    1행: [월] [시작일값] "시작일" 그리고 4열부터 날짜가 **2열씩**
    2행: 노선 | 순번 | 차번 | 오전 | 오후 | 오전 | 오후 …
    3행: 시프트 인덱스(1=오전, 2=오후)
    4행~: 데이터. 노선이 바뀌는 지점 사이에 빈 행 1개
"""
from __future__ import annotations

import datetime as dt
import re
from typing import Any, Optional

from ..models import CellState, DayEntry, DepotGroup, LeaveType, MonthlyRoster
from .weekly import classify_cell

VEHICLE_RE = re.compile(r"^\d{3,4}$")
HEADER_ROUTE = "노선"
HEADER_SLOT = "순번"
HEADER_VEHICLE = "차번"


def _as_date(v: Any) -> Optional[dt.date]:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        m = re.match(r"(\d{4})-(\d{2})-(\d{2})", v.strip())
        if m:
            return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def looks_like_monthly_sheet(ws) -> bool:
    """이 시트가 월간배차 포맷인지 판별 (임포트 시 자동 분기용)."""
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=6):
        labels = [_text(c.value) for c in row]
        if HEADER_ROUTE in labels and HEADER_SLOT in labels and HEADER_VEHICLE in labels:
            return True
    return False


def extract_depot_names(wb) -> dict[str, list[str]]:
    """일일배차/프린터용 시트에서 출발지 이름을 뽑는다 (노선 → 블록 순서).

    월간배차 시트에는 출발지명이 없어 그룹을 ①②로만 구분할 수밖에 없다.
    그런데 담당자가 아는 이름은 '가좌출발/동춘출발'이지 '16번-1'이 아니다.
    게시물이 실물과 같아 보이려면 이 이름을 살려야 한다.

    레이아웃(실측): 2행에 노선 헤더가 가로로, 각 노선 블록 안에서
    '…출발' 라벨이 블록마다 한 번씩 나온다.
    """
    for name in ("프린터용", "일일배차"):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        routes: list[tuple[int, str]] = []   # (열, 노선명)
        labels: list[tuple[int, int, str]] = []  # (행, 열, 출발지)
        for row in ws.iter_rows(max_row=60):
            for c in row:
                v = c.value
                if not isinstance(v, str):
                    continue
                t = v.strip()
                if c.row <= 3 and re.match(r"^\d+(-\d+)?번$", t):
                    routes.append((c.column, t[:-1]))
                elif t.endswith("출발") and len(t) <= 8:
                    labels.append((c.row, c.column, t))
        if not routes or not labels:
            continue
        routes.sort()
        out: dict[str, list[str]] = {}
        for row_i, col, label in sorted(labels):
            # 라벨이 속한 노선 = 시작 열이 라벨보다 작거나 같은 것 중 가장 오른쪽
            owner = None
            for rc, rname in routes:
                if rc <= col:
                    owner = rname
                else:
                    break
            if owner is None:
                continue
            out.setdefault(owner, [])
            if label not in out[owner]:
                out[owner].append(label)
        if out:
            return out
    return {}


def parse_monthly_sheet(
    ws, year: int, month: int, division: str = "",
    depot_names: dict[str, list[str]] | None = None,
) -> MonthlyRoster:
    """월간배차 시트 → MonthlyRoster.

    year/month 는 시트에서 날짜를 읽으므로 검증용으로만 쓰인다.
    """
    grid: dict[tuple[int, int], Any] = {}
    max_col = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                grid[(c.row, c.column)] = c.value
                max_col = max(max_col, c.column)

    # ── 헤더 행 찾기 (노선/순번/차번 라벨이 있는 행) ──
    header_row = None
    col_route = col_slot = col_vehicle = None
    for r in range(1, 8):
        labels = {
            _text(grid.get((r, c))): c for c in range(1, 10)
        }
        if HEADER_ROUTE in labels and HEADER_VEHICLE in labels:
            header_row = r
            col_route = labels[HEADER_ROUTE]
            col_slot = labels.get(HEADER_SLOT)
            col_vehicle = labels[HEADER_VEHICLE]
            break
    if header_row is None:
        raise ValueError("월간배차: '노선/순번/차번' 헤더를 찾지 못했습니다")

    # ── 날짜 열 매핑 ──
    # 헤더 행 위(보통 1행)에서 날짜를 읽고, 오전/오후 라벨과 짝지어
    # (날짜 → 오전열, 오후열) 을 만든다.
    date_row = header_row - 1
    if date_row < 1:
        raise ValueError("월간배차: 날짜 행을 찾지 못했습니다")

    day_cols: dict[dt.date, dict[str, int]] = {}
    for c in range(col_vehicle + 1, max_col + 1):
        d = _as_date(grid.get((date_row, c)))
        if d is None:
            continue
        label = _text(grid.get((header_row, c)))
        if label not in ("오전", "오후"):
            continue
        slot = day_cols.setdefault(d, {})
        # 같은 날짜·같은 라벨이 두 번 나오면 첫 번째만 (병합 셀 대비)
        slot.setdefault("am" if label == "오전" else "pm", c)

    if not day_cols:
        raise ValueError("월간배차: 날짜 열을 찾지 못했습니다")

    roster = MonthlyRoster(year=year, month=month, division=division)
    groups: dict[str, DepotGroup] = {}
    current_route = ""
    depot_idx = 0
    prev_slot = 0

    # ── 데이터 행 ──
    r = header_row + 1
    blanks = 0
    max_row = max(rr for rr, _ in grid)
    while r <= max_row and blanks < 25:
        vehicle = _text(grid.get((r, col_vehicle)))
        if not VEHICLE_RE.match(vehicle):
            blanks += 1
            r += 1
            continue
        blanks = 0

        route = _text(grid.get((r, col_route)))
        # 노선 셀이 매 행에 반복될 수도, 블록 첫 행에만 있을 수도 있다.
        # '값이 있을 때'가 아니라 '노선이 바뀔 때'만 그룹 상태를 초기화한다.
        if route and route != current_route:
            current_route = route
            depot_idx = 0
            prev_slot = 0

        slot_raw = _text(grid.get((r, col_slot))) if col_slot else ""
        slot_index = int(slot_raw) if slot_raw.isdigit() else None

        # 한 노선 안에서 순번이 1로 되돌아가면 다른 출발지그룹이다.
        # (실측: 16번 = 가좌출발 7대 + 동춘출발 7대 → 순번 1~7 이 두 번 반복)
        # 로테이션은 출발지그룹 안에서만 돌기 때문에 반드시 분리해야 한다.
        if slot_index is not None and slot_index <= prev_slot:
            depot_idx += 1
        if slot_index is not None:
            prev_slot = slot_index

        route_name = f"{current_route}번" if current_route else "전체"
        # 출발지 이름을 알면 그대로 쓴다 (담당자가 아는 말: '가좌출발').
        # 모르면 번호로 구분한다.
        known = (depot_names or {}).get(current_route, [])
        if depot_idx < len(known):
            suffix = known[depot_idx]
        else:
            circled = "①②③④⑤⑥⑦⑧⑨"
            suffix = circled[depot_idx] if depot_idx < len(circled) else str(depot_idx + 1)
        group_name = f"{route_name} {suffix}"

        g = groups.get(group_name)
        if g is None:
            g = DepotGroup(name=group_name)
            groups[group_name] = g
            roster.groups.append(g)
        if vehicle not in g.vehicles:
            g.vehicles.append(vehicle)

        for d, cols in day_cols.items():
            am = classify_cell(grid.get((r, cols["am"]))) if "am" in cols else classify_cell(None)
            pm = classify_cell(grid.get((r, cols["pm"]))) if "pm" in cols else classify_cell(None)
            # 이 양식은 휴차를 따로 표기하지 않는다 — 오전·오후가 모두 비면
            # 그 차량은 그날 운행하지 않은 것으로 본다 (주간배차표의 '휴'와 동일).
            # 그래야 하류의 감차 패턴 추론이 두 양식에서 같게 동작한다.
            if not am.driver and not am.raw and not pm.driver and not pm.raw:
                am = CellState(leave=LeaveType.REGULAR, raw="휴")
                pm = CellState(leave=LeaveType.REGULAR, raw="휴")
            roster.entries[(d, vehicle)] = DayEntry(
                date=d, vehicle=vehicle,
                slot_label=slot_raw or None,
                slot_index=slot_index,
                am=am, pm=pm,
            )
        r += 1

    # 출발지그룹이 하나뿐인 노선은 접미사를 떼어 이름을 깔끔하게
    from collections import Counter as _Counter

    base_count = _Counter(g.name.rsplit(" ", 1)[0] for g in roster.groups)
    for g in roster.groups:
        base = g.name.rsplit(" ", 1)[0]
        if base_count[base] == 1:
            g.name = base

    return roster


def parse_monthly_workbook(
    path: str, sheet_name: str = "월간배차", division: str = ""
) -> MonthlyRoster:
    """작업용 워크북에서 월간배차 시트를 읽는다.

    연·월은 시트의 날짜에서 직접 판별한다 (파일명이 제각각이라 신뢰 불가).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        # 연·월 선판별: 날짜 행에서 가장 많이 등장하는 (연, 월)
        from collections import Counter

        ym: Counter = Counter()
        for row in ws.iter_rows(min_row=1, max_row=3):
            for c in row:
                d = _as_date(c.value)
                if d:
                    ym[(d.year, d.month)] += 1
        if not ym:
            raise ValueError("월간배차: 날짜를 찾지 못했습니다")
        (year, month), _ = ym.most_common(1)[0]
        return parse_monthly_sheet(
            ws, year, month, division=division,
            depot_names=extract_depot_names(wb),
        )
    finally:
        wb.close()
