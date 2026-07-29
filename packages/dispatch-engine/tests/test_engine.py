"""엔진 핵심 검증 테스트.

실데이터 워크북(로컬 Downloads)이 있으면 백테스트 불변식을 검증하고,
없으면 합성 데이터로 로테이션·솔버 스모크 테스트만 수행한다.
실행: python -m pytest tests/ -x -q
"""
import datetime as dt
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from busync_engine.config import (
    DayClass,
    DisplayMode,
    GroupReductionConfig,
    ReductionCalendar,
    ReductionMode,
)
from busync_engine.importer.inference import RotationRule
from busync_engine.rotation import expand_pattern
from busync_engine.solver import AssignmentProblem, solve

JISEON = os.path.expanduser("~/Downloads/2026년 지선배차.xlsx")
GANSEON = os.path.expanduser("~/Downloads/2026년 간선배차.xlsx")


def test_rotation_step2_cycle():
    """5대 그룹 +2 로테이션: 5일 주기로 원위치, 월 경계 연속."""
    rule = RotationRule(
        group="가좌", size=5,
        perm={1: 3, 2: 4, 3: 5, 4: 1, 5: 2}, support=1.0,
    )
    last = {"V1": 1, "V2": 2, "V3": 3, "V4": 4, "V5": 5}
    cal = ReductionCalendar()
    pat = expand_pattern(
        rule, last, dt.date(2026, 7, 1), dt.date(2026, 7, 10), cal
    )
    assert pat[(dt.date(2026, 7, 1), "V1")].underlying_slot == 3
    # 5일 주기 복귀
    assert (
        pat[(dt.date(2026, 7, 5), "V1")].underlying_slot
        == pat[(dt.date(2026, 7, 10), "V1")].underlying_slot
    )


def test_reduction_fixed_slots_overlay():
    """고정 휴차 슬롯: 주말에 해당 슬롯 차량만 미운행, 로테이션은 계속."""
    rule = RotationRule(
        group="G", size=5, perm={1: 3, 2: 4, 3: 5, 4: 1, 5: 2}, support=1.0
    )
    last = {"V1": 1, "V2": 2, "V3": 3, "V4": 4, "V5": 5}
    cal = ReductionCalendar()
    cfg = GroupReductionConfig(
        mode=ReductionMode.FIXED_SLOTS,
        rest_slots={DayClass.SAT: frozenset({3}), DayClass.SUNHOL: frozenset({3})},
    )
    # 2026-07-04는 토요일
    pat = expand_pattern(
        rule, last, dt.date(2026, 7, 1), dt.date(2026, 7, 5), cal, cfg,
        DisplayMode.KEEP,
    )
    sat = dt.date(2026, 7, 4)
    resting = [v for v in last if not pat[(sat, v)].operating]
    assert len(resting) == 1
    assert pat[(sat, resting[0])].underlying_slot == 3
    # KEEP 모드: 휴차 차량도 표시 순번 유지
    assert pat[(sat, resting[0])].display_slot == 3


def test_reduction_compact_display():
    """COMPACT 모드(간선): 운행 차량만 1..M 재부여."""
    rule = RotationRule(
        group="G", size=5, perm={1: 5, 2: 1, 3: 2, 4: 3, 5: 4}, support=1.0
    )  # -1 로테이션
    last = {"V1": 1, "V2": 2, "V3": 3, "V4": 4, "V5": 5}
    cal = ReductionCalendar()
    cfg = GroupReductionConfig(
        mode=ReductionMode.FIXED_SLOTS,
        rest_slots={DayClass.SAT: frozenset({4, 5}), DayClass.SUNHOL: frozenset()},
    )
    pat = expand_pattern(
        rule, last, dt.date(2026, 7, 4), dt.date(2026, 7, 4), cal, cfg,
        DisplayMode.COMPACT,
    )
    sat = dt.date(2026, 7, 4)
    ops = sorted(
        (c.display_slot for c in (pat[(sat, v)] for v in last) if c.operating)
    )
    assert ops == [1, 2, 3]
    assert all(
        pat[(sat, v)].display_slot is None
        for v in last if not pat[(sat, v)].operating
    )


def test_solver_smoke_h1_h2():
    """미니 문제: H1(슬롯당 1명)·H2(1일 1시프트)·S1(본인차량) 충족."""
    days = [dt.date(2026, 7, 1), dt.date(2026, 7, 2)]
    operating = {(d, v, s) for d in days for v in ("V1", "V2") for s in ("A", "P")}
    problem = AssignmentProblem(
        dates=days,
        operating=operating,
        drivers=["a", "b", "c", "d"],
        home_vehicle={"a": "V1", "b": "V1", "c": "V2", "d": "V2"},
        partner={"a": "b", "b": "a", "c": "d", "d": "c"},
        work_days_band=(2, 2),
        hard_own_vehicle=True,
    )
    asg = solve(problem, time_limit_s=10)
    assert len(asg.cells) == 8
    for d in days:
        workers = [asg.cells[(d, v, s)] for v in ("V1", "V2") for s in ("A", "P")]
        assert len(set(workers)) == 4  # H2
        assert asg.cells[(d, "V1", "A")] in ("a", "b")  # S1 하드


@pytest.mark.skipif(not os.path.exists(JISEON), reason="실데이터 없음")
def test_backtest_stage1_jiseon_may_2026_is_perfect():
    from busync_engine.backtest import backtest_stage1
    from busync_engine.importer.weekly import parse_workbook_month

    prev = parse_workbook_month(JISEON, "지선배차표(2026년 4월)", "지선")
    tgt = parse_workbook_month(JISEON, "지선배차표(2026년 5월)", "지선")
    holidays = {dt.date(2026, 5, 1), dt.date(2026, 5, 5), dt.date(2026, 5, 25)}
    res, _ = backtest_stage1(prev, tgt, holidays)
    assert res.slot_rate == 1.0


@pytest.mark.skipif(not os.path.exists(GANSEON), reason="실데이터 없음")
def test_backtest_stage1_ganseon_may_2026_is_perfect():
    from busync_engine.backtest import backtest_stage1
    from busync_engine.importer.weekly import parse_workbook_month

    prev = parse_workbook_month(GANSEON, "간선2026년4월", "간선")
    tgt = parse_workbook_month(GANSEON, "간선2026년5월", "간선")
    holidays = {dt.date(2026, 5, 1), dt.date(2026, 5, 5), dt.date(2026, 5, 25)}
    res, _ = backtest_stage1(prev, tgt, holidays)
    assert res.slot_rate == 1.0


def test_policy_catalog_integrity():
    from busync_engine.policy import (
        CATALOG_BY_KEY, CompanyPolicy, SETTINGS_CATALOG, catalog_as_json,
    )

    keys = [s.key for s in SETTINGS_CATALOG]
    assert len(keys) == len(set(keys)), "설정 키 중복"
    assert len(catalog_as_json()) == len(SETTINGS_CATALOG)
    p = CompanyPolicy()
    # 모든 키에 기본값 존재
    for k in keys:
        assert p.get(k) is not None or CATALOG_BY_KEY[k].default is None
    # 직렬화 라운드트립
    p.set("max_consecutive_days", 5)
    p.holidays = {dt.date(2026, 5, 5)}
    p2 = CompanyPolicy.from_dict(p.to_dict())
    assert p2.get("max_consecutive_days") == 5
    assert p2.holidays == {dt.date(2026, 5, 5)}
    with pytest.raises(KeyError):
        p.set("없는키", 1)


@pytest.mark.skipif(not os.path.exists(JISEON), reason="실데이터 없음")
def test_recommend_on_real_data():
    from busync_engine.importer.weekly import parse_workbook_month
    from busync_engine.policy import CATALOG_BY_KEY
    from busync_engine.recommend import analyze

    rosters = [
        parse_workbook_month(JISEON, f"지선배차표(2026년 {m}월)", "지선")
        for m in (4, 5)
    ]
    rep = analyze(rosters)
    assert rep.drivers > 30
    assert rep.group_rules and all(
        g.rotation_support > 0.9 for g in rep.group_rules
    )
    # 추천 키는 전부 카탈로그에 존재해야 함
    for r in rep.recommendations:
        assert r.key in CATALOG_BY_KEY, r.key
        assert 0 <= r.confidence <= 1
        assert r.evidence


def test_renderer_roundtrip_synthetic(tmp_path):
    """렌더러 산출물을 자체 임포터로 재파싱했을 때 동일해야 한다."""
    from busync_engine.importer.weekly import parse_weekly_sheet
    from busync_engine.models import (
        CellState, DayEntry, DepotGroup, MonthlyRoster,
    )
    from busync_engine.renderer import render_weekly_xlsx
    import openpyxl

    roster = MonthlyRoster(year=2026, month=7, division="지선")
    g = DepotGroup(name="테스트", vehicles=["1001", "1002"])
    roster.groups.append(g)
    d0 = dt.date(2026, 7, 1)
    for i in range(31):
        d = d0 + dt.timedelta(days=i)
        for vi, v in enumerate(g.vehicles):
            slot = (i + vi) % 2 + 1
            roster.entries[(d, v)] = DayEntry(
                date=d, vehicle=v, slot_label=str(slot), slot_index=slot,
                am=CellState(driver=f"기사{vi}A", raw=f"기사{vi}A"),
                pm=CellState(driver=f"기사{vi}P", raw=f"기사{vi}P"),
            )
    out = tmp_path / "render.xlsx"
    render_weekly_xlsx(roster, str(out))
    wb = openpyxl.load_workbook(str(out))
    r2 = parse_weekly_sheet(wb.active, 2026, 7, "지선")
    assert len(r2.month_dates()) == 31
    e = r2.entry(dt.date(2026, 7, 15), "1001")
    assert e.am.driver == "기사0A" and e.pm.driver == "기사0P"


def _mini_generation():
    """설명·수리 테스트 공용 미니 생성 문제."""
    from busync_engine.solver import AssignmentProblem, solve

    days = [dt.date(2026, 7, i) for i in range(1, 8)]
    operating = {
        (d, v, s) for d in days for v in ("V1", "V2") for s in ("A", "P")
    }
    problem = AssignmentProblem(
        dates=days,
        operating=operating,
        drivers=["a", "b", "c", "d", "e"],
        leaves={"a": {dt.date(2026, 7, 4)}, "b": {dt.date(2026, 7, 4)}},
        home_vehicle={"a": "V1", "b": "V1", "c": "V2", "d": "V2"},
        partner={"a": "b", "b": "a", "c": "d", "d": "c"},
        work_days_band=(4, 7),
        hard_own_vehicle=True,
        allow_unfilled=True,
    )
    asg = solve(problem, time_limit_s=15)
    return problem, asg


def test_explain_cell_gives_reasons():
    from busync_engine.explain import explain_cell, explain_driver_month

    problem, asg = _mini_generation()
    d = dt.date(2026, 7, 2)
    driver = asg.cells[(d, "V1", "A")]
    exp = explain_cell(problem, asg, d, "V1", "A")
    assert exp.driver == driver
    codes = {r.code for r in exp.reasons}
    assert "OWN_VEHICLE" in codes or "SPARE" in codes
    assert exp.summary.startswith(driver)
    lines = explain_driver_month(problem, asg, driver)
    assert any("근무" in ln for ln in lines)


def test_repair_flow():
    from busync_engine.repair import (
        apply_repair, changed_cells, mark_absent, suggest_repair,
    )

    problem, asg = _mini_generation()
    d = dt.date(2026, 7, 3)
    key = (d, "V1", "A")
    absent = asg.cells[key]
    before = dict(asg.cells)
    assert mark_absent(asg, d, "V1", "A") == absent
    sug = suggest_repair(problem, asg, d, "V1", "A")
    # 결원자 본인은 후보에서 제외
    assert all(c.driver != absent for c in sug.candidates)
    assert sug.absent_driver == absent
    assert sug.top, "가용 후보가 있어야 함"
    rec = apply_repair(asg, d, "V1", "A", sug.top[0].driver, reason="병가")
    assert rec.removed == absent and rec.added == sug.top[0].driver
    diff = changed_cells(before, asg.cells)
    assert len(diff) == 1 and diff[0][:3] == (d, "V1", "A")


def test_annual_leave_labor_law():
    from busync_engine.leave import annual_leave_days, prorated_work_days

    as_of = dt.date(2026, 7, 27)
    assert annual_leave_days(dt.date(2026, 1, 15), as_of) == 6   # 1년 미만 월 1개
    assert annual_leave_days(dt.date(2025, 3, 1), as_of) == 15   # 1년 이상
    assert annual_leave_days(dt.date(2020, 3, 1), as_of) == 17   # 6년차 +2
    assert annual_leave_days(dt.date(2005, 1, 1), as_of) == 25   # 상한
    assert annual_leave_days(dt.date(2027, 1, 1), as_of) == 0    # 입사 전
    # 월중 입사 일할: 6/14 입사, 만근 22일 기준
    assert prorated_work_days(
        22, dt.date(2026, 6, 1), dt.date(2026, 6, 30),
        hire_date=dt.date(2026, 6, 14),
    ) == 12


def test_leave_triage_priority():
    from busync_engine.leave import (
        LeaveRequest, RequestStatus, SubstituteCounter, triage,
    )
    from busync_engine.models import LeaveType

    now = dt.datetime(2026, 8, 1, 9, 0)
    d = dt.date(2026, 8, 15)
    counters = {
        "받아준사람": SubstituteCounter(requested=0, accepted=5),
        "신세진사람": SubstituteCounter(requested=5, accepted=0),
    }
    reqs = [
        LeaveRequest("받아준사람", d, requested_at=now + dt.timedelta(hours=1)),
        LeaveRequest("신세진사람", d, requested_at=now),
        LeaveRequest("연차사람", d, leave_type=LeaveType.ANNUAL,
                     requested_at=now + dt.timedelta(hours=2)),
    ]
    decisions = triage(reqs, counters, daily_cap=2)
    dec = decisions[0]
    approved = {r.driver for r in dec.auto_approved}
    # 연차는 법정 권리로 최우선, 호혜성 높은 사람이 다음
    assert "연차사람" in approved
    assert "받아준사람" in approved
    assert dec.needs_review[0].request.driver == "신세진사람"
    assert dec.needs_review[0].request.status == RequestStatus.PENDING
    # 정원 이내면 전원 자동 승낙
    reqs2 = [LeaveRequest("한명", d)]
    dec2 = triage(reqs2, {}, daily_cap=6)[0]
    assert dec2.auto_approved[0].status == RequestStatus.APPROVED


WORKBOOK_ZIP_DIR = (
    "/private/tmp/claude-501/-Users-tompark-Desktop-busync/"
    "7d70804a-9697-4f1e-a9da-50964619dd1b/scratchpad/data/ganseon_zip"
)
WORK_FILE = os.path.join(WORKBOOK_ZIP_DIR, "배차표(2026 .5) 배차 작업용.xlsx")


@pytest.mark.skipif(not os.path.exists(WORK_FILE), reason="작업용 실데이터 없음")
def test_overview_parser_matches_reported_workdays():
    """배차총괄: 파싱한 A/P 카운트가 시트 기재 근무일수와 전원 일치해야 한다."""
    import openpyxl
    from busync_engine.importer.worksheets import parse_overview

    wb = openpyxl.load_workbook(WORK_FILE, read_only=True, data_only=True)
    ov = parse_overview(wb["배차총괄"])
    wb.close()
    assert ov.year == 2026 and ov.month == 5
    assert len(ov.drivers) > 100
    mismatch = [
        r.name for r in ov.drivers
        if r.work_days_reported is not None
        and r.work_days_reported != r.work_days_counted
    ]
    assert mismatch == []


@pytest.mark.skipif(not os.path.exists(WORK_FILE), reason="작업용 실데이터 없음")
def test_daily_parser_routes_and_vacancies():
    """일일배차: 3개 노선 × 2개 출발지 블록, 결원(0) 슬롯 구분."""
    import openpyxl
    from busync_engine.importer.worksheets import parse_daily

    wb = openpyxl.load_workbook(WORK_FILE, read_only=True, data_only=True)
    daily = parse_daily(wb["일일배차"])
    wb.close()
    assert daily.date is not None
    assert set(daily.routes()) == {"16", "9", "3-2"}
    depots = {s.depot for s in daily.slots}
    assert any("가좌" in d for d in depots) and any("동춘" in d for d in depots)
    # 조율 열 = 로테이션 오프셋 가시화 (스펙 3.2)
    assert any(s.adjust is not None for s in daily.slots)
    # 결원 슬롯 존재 + 기사명이 "0"으로 오염되지 않음
    assert all(s.am != "0" and s.pm != "0" for s in daily.slots)


@pytest.mark.skipif(not os.path.exists(JISEON), reason="실데이터 없음")
def test_generate_requires_immediately_previous_month():
    """로테이션은 전월 말일에서 이어받으므로 직전 월 이력이 없으면 거부해야 한다.

    (예전엔 replay_underlying에서 KeyError로 터졌다 — 원인을 알 수 없는 실패)
    """
    from busync_engine.generate import generate_month
    from busync_engine.importer.weekly import parse_workbook_month
    from busync_engine.policy import CompanyPolicy

    history = [
        parse_workbook_month(JISEON, "지선배차표(2026년 5월)", "지선"),
        parse_workbook_month(JISEON, "지선배차표(2026년 6월)", "지선"),
    ]
    # 이력은 6월까지인데 8월을 요청 → 7월이 비어 로테이션을 이을 수 없다
    with pytest.raises(ValueError, match="직전 월"):
        generate_month(history, CompanyPolicy(), 2026, 8, time_limit_s=5)

    # 다만 '월 경계 이어가기'를 끄면 직전 월 없이도 생성돼야 한다
    # (첫 도입처럼 지난달 자료가 아예 없는 경우)
    policy = CompanyPolicy()
    policy.set("rotation_carry_over", False)
    result = generate_month(history, policy, 2026, 8, time_limit_s=20)
    assert result.roster.entries, "이어받기를 꺼도 배차표는 생성되어야 함"
    assert any("새로 시작" in w for w in result.warnings)
