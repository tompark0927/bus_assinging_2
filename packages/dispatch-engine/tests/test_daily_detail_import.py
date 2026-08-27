"""Busync 가 내보낸 배차표를 다시 올리는 경로 회귀 테스트.

실제 사고: 담당자가 7월 배차표를 [Excel 내보내기]로 받아 8월 생성에 올렸는데
"차량번호가 하나도 일치하지 않습니다"가 떴다. 원인은 기초 데이터가 아니라,
내보낸 파일의 첫 시트가 행=기사/칸=노선번호라 차량 중심 파서로 한 칸도
읽히지 않은 것. 그때 엔진이 에러 없이 빈 결과를 돌려준 게 진짜 문제였다.
"""
import datetime as dt

import openpyxl
import pytest

from busync_engine.importer.daily_detail import has_daily_detail, parse_daily_detail


def _make_export(tmp_path, rows, sheet_first="2026년 7월 배차표"):
    """내보내기 파일 모양 재현 — 첫 시트(기사 중심) + '일별 상세'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_first
    ws.append(["성민버스 2026년 7월 배차표"])
    ws.append(["사원번호", "이름", "구분", "1\n(수)"])
    ws.append(["DRV062", "강구인", "메인", "16"])

    d = wb.create_sheet("일별 상세")
    d.append(["날짜", "요일", "노선", "기사 이름", "사원번호", "구분", "버스번호", "근무형태", "상태"])
    for r in rows:
        d.append(r)
    p = tmp_path / "export.xlsx"
    wb.save(p)
    return p


def test_내보낸_파일의_일별상세를_읽는다(tmp_path):
    p = _make_export(tmp_path, [
        ["2026.07.01", "수", "16", "김명천", "DRV037", "메인", "2506", "오전", "정상"],
        ["2026.07.01", "수", "16", "김상윤", "DRV038", "메인", "2506", "오후", "정상"],
        ["2026.07.02", "목", "9", "김성재", "DRV039", "메인", "2508", "오전", "정상"],
    ])
    wb = openpyxl.load_workbook(p, data_only=True)
    assert has_daily_detail(wb)

    roster = parse_daily_detail(wb)
    assert roster is not None
    assert (roster.year, roster.month) == (2026, 7)

    e = roster.entry(dt.date(2026, 7, 1), "2506")
    assert e is not None
    # 같은 차량·같은 날의 오전/오후가 한 칸으로 합쳐진다
    assert e.am.driver == "김명천"
    assert e.pm.driver == "김상윤"
    assert roster.drivers() == {"김명천", "김상윤", "김성재"}


def test_종일_근무는_오전오후_모두_채운다(tmp_path):
    p = _make_export(tmp_path, [
        ["2026.07.01", "수", "16", "한재익", "DRV001", "메인", "2510", "종일", "정상"],
    ])
    roster = parse_daily_detail(openpyxl.load_workbook(p, data_only=True))
    e = roster.entry(dt.date(2026, 7, 1), "2510")
    assert e.am.driver == "한재익" and e.pm.driver == "한재익"


def test_휴무_표기는_기사로_세지_않는다(tmp_path):
    p = _make_export(tmp_path, [
        ["2026.07.01", "수", "16", "휴", "", "메인", "2506", "오전", "정상"],
        ["2026.07.01", "수", "16", "김상윤", "DRV038", "메인", "2506", "오후", "정상"],
    ])
    roster = parse_daily_detail(openpyxl.load_workbook(p, data_only=True))
    e = roster.entry(dt.date(2026, 7, 1), "2506")
    assert e.am.driver is None
    assert roster.drivers() == {"김상윤"}


def test_일별상세가_없으면_None(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "2026년 7월 배차표"
    p = tmp_path / "no_detail.xlsx"
    wb.save(p)
    loaded = openpyxl.load_workbook(p, data_only=True)
    assert not has_daily_detail(loaded)
    assert parse_daily_detail(loaded) is None


def test_import_엔드포인트가_내보낸_파일을_읽는다(tmp_path):
    from fastapi.testclient import TestClient
    import service

    p = _make_export(tmp_path, [
        ["2026.07.01", "수", "16", "김명천", "DRV037", "메인", "2506", "오전", "정상"],
        ["2026.07.01", "수", "16", "김상윤", "DRV038", "메인", "2506", "오후", "정상"],
    ])
    client = TestClient(service.app)
    with open(p, "rb") as f:
        r = client.post("/import", files={"file": ("export.xlsx", f, "application/vnd.ms-excel")})
    assert r.status_code == 200, r.text
    body = r.json()
    # 여기가 사고의 핵심 — 예전에는 전부 0 이 나가고 백엔드가 기초 데이터를 의심했다
    assert body["vehicles"] == 1
    assert body["filled_cells"] == 2
    assert body["cells"]["2026-07-01"]["2506"]["am"] == "김명천"


def test_한_칸도_못_읽으면_조용히_빈결과를_주지_않는다(tmp_path):
    """빈 결과가 나가면 백엔드가 '차량번호가 하나도 일치하지 않습니다'로 오도한다."""
    from fastapi.testclient import TestClient
    import service

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026년 7월 배차표"
    ws.append(["사원번호", "이름", "구분", "1\n(수)"])
    ws.append(["DRV062", "강구인", "메인", "16"])
    p = tmp_path / "driver_view_only.xlsx"
    wb.save(p)

    client = TestClient(service.app)
    with open(p, "rb") as f:
        r = client.post("/import", files={"file": ("x.xlsx", f, "application/vnd.ms-excel")})
    assert r.status_code == 422
    assert "한 칸도 읽지 못했습니다" in r.json()["detail"]


def test_순번이_없는_파일이면_이어받는_척하지_않고_새로_시작한다(tmp_path):
    """이어받을 순번이 없는데 전월 말일에서 이어받으면 그게 곧 틀린 순번이다.

    막아버리면 담당자가 8월 배차를 아예 못 짠다(지난달 배차표가 Busync 안에
    있는데도). 새로 시작하되, 이어지지 않는다는 사실을 경고로 분명히 남긴다.
    """
    from fastapi.testclient import TestClient
    import service

    # 실제 성민버스처럼 '등록 3대 중 매일 2대만 운행'을 만든다. 감차가 있으면
    # 전 차량이 순번을 가진 날이 하루도 없어 시트에서 회전을 읽을 수 없다.
    fleet = ["2506", "2507", "2508"]
    rows = []
    for day in range(1, 32):
        running = [v for i, v in enumerate(fleet) if i != day % 3]  # 매일 한 대씩 감차
        for v in running:
            rows.append([f"2026.07.{day:02d}", "수", "16", f"기사{v}", "D1", "메인", v, "오전", "정상"])
            rows.append([f"2026.07.{day:02d}", "수", "16", f"부기사{v}", "D2", "메인", v, "오후", "정상"])
    p = _make_export(tmp_path, rows)

    client = TestClient(service.app)
    with open(p, "rb") as f:
        r = client.post(
            "/generate",
            files={"file": ("export.xlsx", f, "application/vnd.ms-excel")},
            data={"year": "2026", "month": "8"},
        )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["cells"], "배차표가 생성되어야 한다"

    # 순번이 이어지지 않는다는 사실을 반드시 알린다 — 조용히 새로 시작하면
    # 담당자는 지난달과 이어진 줄 알고 그대로 게시한다
    warns = " ".join(body.get("warnings") or [])
    assert "순번 정보가 없어" in warns
    assert "이어지지 않" in warns
