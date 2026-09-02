"""BuSync 배차 엔진 HTTP 서비스 (FastAPI).

Node 백엔드(Express)와의 연동용 마이크로서비스. 솔버(OR-Tools)가 Python
전용이라 엔진은 별도 프로세스로 뜨고, 백엔드가 프록시한다.

실행: uvicorn service:app --port 8100

엔드포인트
  GET  /health              상태 확인
  GET  /catalog             설정 카탈로그 (프론트 설정 화면 렌더링용)
  POST /analyze             엑셀 업로드 → 규칙 감지 + 설정 추천 (온보딩 위저드)
  POST /backtest            과거 월 재현 검증 리포트
  POST /generate            월 배차 초안 생성 → draft_id 반환 (+감사/공정성/xlsx)
  GET  /draft/{id}/explain  셀 배정 근거 ("왜 이 기사가 이 슬롯인가")
  GET  /draft/{id}/explain-driver  기사 월간 설명 (기사 앱)
  POST /draft/{id}/absence  당일 결원 신고 → 대체 후보 3명 추천
  POST /draft/{id}/repair   대체 확정 (해당 셀만 교체 + 변경 마킹)
  POST /leave/triage        휴무신청 자동 승낙/검토 분류 (호혜성 스코어)
  GET  /leave/annual        근로기준법 연차 자동계산

초안(draft)은 프로세스 메모리에 보관된다(최근 20건). 영속화·권한은 Node
백엔드 소관 — 엔진은 계산만 담당한다.
"""
from __future__ import annotations

import base64
import datetime as dt
import json
import os
import pickle
import tempfile
import uuid
from collections import OrderedDict, defaultdict
from pathlib import Path

from fastapi import FastAPI, File, Form, Header, HTTPException, UploadFile

from busync_engine.audit import audit as run_audit
from busync_engine.backtest import backtest_stage1, backtest_stage2
from busync_engine.generate import generate_month
from busync_engine.inspector import inspect_roster
from busync_engine.importer.monthly import (
    extract_depot_names,
    looks_like_monthly_sheet,
    parse_monthly_sheet,
)
from busync_engine.importer.daily_detail import has_daily_detail, parse_daily_detail
from busync_engine.importer.from_cells import roster_from_cells
from busync_engine.importer.weekly import SHEET_YM_RE, parse_workbook_month
from busync_engine.policy import CompanyPolicy, catalog_as_json
from busync_engine.recommend import analyze

app = FastAPI(title="BuSync Dispatch Engine", version="0.4.0")

# 데이터 디렉토리 (정책·초안 영속화). ENGINE_DATA_DIR로 재지정 가능.
DATA_DIR = Path(os.environ.get("ENGINE_DATA_DIR", Path(__file__).parent / "data"))
POLICY_DIR = DATA_DIR / "policies"
DRAFT_DIR = DATA_DIR / "drafts"
for _d in (POLICY_DIR, DRAFT_DIR):
    _d.mkdir(parents=True, exist_ok=True)

# 초안 캐시: draft_id -> {result, repairs}. 디스크에도 pickle로 영속화되어
# 프로세스 재시작 후에도 접근 가능 (최근 것부터 메모리 20건 캐시).
_DRAFTS: OrderedDict[str, dict] = OrderedDict()
_DRAFTS_MAX = 20


def _draft_path(draft_id: str) -> Path:
    if not draft_id.isalnum():
        raise HTTPException(400, "잘못된 draft_id")
    return DRAFT_DIR / f"{draft_id}.pkl"


def _persist_draft(draft_id: str) -> None:
    with open(_draft_path(draft_id), "wb") as f:
        pickle.dump(_DRAFTS[draft_id], f)


def _store_draft(result) -> str:
    draft_id = uuid.uuid4().hex[:12]
    _DRAFTS[draft_id] = {"result": result, "repairs": []}
    _persist_draft(draft_id)
    while len(_DRAFTS) > _DRAFTS_MAX:
        _DRAFTS.popitem(last=False)
    return draft_id


def _get_draft(draft_id: str) -> dict:
    if draft_id not in _DRAFTS:
        path = _draft_path(draft_id)
        if not path.exists():
            raise HTTPException(404, "초안을 찾을 수 없습니다 (만료되었을 수 있음)")
        with open(path, "rb") as f:
            _DRAFTS[draft_id] = pickle.load(f)
    return _DRAFTS[draft_id]


def _company_policy_path(company_id: str) -> Path:
    safe = "".join(c for c in company_id if c.isalnum() or c in "-_") or "default"
    return POLICY_DIR / f"{safe}.json"


def _save_upload(file: UploadFile) -> str:
    suffix = ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
        f.write(file.file.read())
        return f.name


def _load_rosters(path: str, division: str, sheets: list[str]) -> list:
    """시트 이름 목록 → MonthlyRoster 목록. 양식은 자동 판별한다.

    - 주간배차표(행=차량, 열=날짜별 순번|오전|오후): 지선·간선 게시용
    - 월간배차(행=노선|순번|차번, 열=날짜별 오전|오후): 담당자 작업용
    두 양식이 담는 정보는 같아서 어느 쪽이든 생성 입력으로 쓸 수 있다.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # Busync 가 내보낸 파일이면 '일별 상세' 시트를 우선한다.
        # 그 파일의 첫 시트는 행=기사 / 셀=노선번호라 아래 차량 중심 파서로는
        # 한 글자도 안 읽힌다 — 그러면 빈 결과가 나가고 백엔드가 애먼 기초
        # 데이터를 의심하게 만든다. '일별 상세' 는 (날짜×차량×오전/오후×기사)를
        # 한 행씩 갖고 있어 오히려 원본 양식보다 정확하다.
        if has_daily_detail(wb):
            roster = parse_daily_detail(wb, division)
            if roster is not None:
                return [roster]

        out = []
        for name in sheets:
            ws = wb[name]
            if looks_like_monthly_sheet(ws):
                from collections import Counter as _C
                import datetime as _dt

                ym: _C = _C()
                for row in ws.iter_rows(min_row=1, max_row=3):
                    for c in row:
                        v = c.value
                        if isinstance(v, (_dt.datetime, _dt.date)):
                            d = v.date() if isinstance(v, _dt.datetime) else v
                            ym[(d.year, d.month)] += 1
                if not ym:
                    raise HTTPException(422, f"'{name}' 시트에서 날짜를 찾지 못했습니다")
                (y, m), _ = ym.most_common(1)[0]
                # read_only 워크시트는 재순회가 필요해 새로 연다
                wb2 = openpyxl.load_workbook(path, data_only=True)
                try:
                    out.append(parse_monthly_sheet(
                        wb2[name], y, m, division,
                        depot_names=extract_depot_names(wb2),
                    ))
                finally:
                    wb2.close()
            else:
                out.append(parse_workbook_month(path, name, division))
        return out
    finally:
        wb.close()


def _month_sheets(path: str, division_hint: str = "") -> list[str]:
    """워크북에서 월 시트 자동 감지 (연·월 파싱 가능한 시트, 시간순)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        found = []
        for name in wb.sheetnames:
            if "패턴" in name:
                continue
            m = SHEET_YM_RE.search(name)
            if m:
                found.append((int(m.group(1)), int(m.group(2)), name))
        if found:
            found.sort()
            return [name for _, _, name in found]
        # 시트명에 연월이 없는 작업용 워크북 — 월간배차 시트를 쓴다
        for name in wb.sheetnames:
            if looks_like_monthly_sheet(wb[name]):
                return [name]
        return []
    finally:
        wb.close()


@app.get("/health")
def health():
    return {"ok": True}


@app.get("/catalog")
def catalog():
    return {"settings": catalog_as_json()}


@app.get("/policy")
def get_policy(x_company_id: str = Header("default")):
    """회사 정책 조회. 미저장 시 카탈로그 기본값으로 빈 정책 반환."""
    path = _company_policy_path(x_company_id)
    if path.exists():
        data = json.loads(path.read_text())
        return {"policy": data, "is_default": False}
    return {"policy": CompanyPolicy().to_dict(), "is_default": True}


@app.put("/policy")
def put_policy(payload: dict, x_company_id: str = Header("default")):
    """회사 정책 저장. 카탈로그에 없는 키는 거부 (오타·스키마 드리프트 방지)."""
    try:
        policy = CompanyPolicy.from_dict(payload.get("policy", payload))
        for key in policy.values:
            CompanyPolicy().get(key)  # 카탈로그 존재 검증 (KeyError → 422)
    except (KeyError, ValueError) as ex:
        raise HTTPException(422, f"정책 형식 오류: {ex}")
    _company_policy_path(x_company_id).write_text(
        json.dumps(policy.to_dict(), ensure_ascii=False, indent=1)
    )
    return {"ok": True, "policy": policy.to_dict()}


@app.post("/analyze")
async def analyze_endpoint(
    file: UploadFile = File(...),
    division: str = Form(""),
    sheets: str = Form(""),          # 쉼표 구분, 비우면 최근 3개월 자동
    months: int = Form(3),
):
    path = _save_upload(file)
    names = [s.strip() for s in sheets.split(",") if s.strip()] or \
        _month_sheets(path)[-months:]
    if not names:
        raise HTTPException(400, "분석할 월 시트를 찾지 못했습니다")
    try:
        rosters = _load_rosters(path, division, names)
    except HTTPException:
        raise
    except Exception as ex:
        raise HTTPException(422, f"파싱 실패: {ex}")
    rep = analyze(rosters)
    return {
        "sheets_analyzed": names,
        "months": rep.months_analyzed,
        "drivers": rep.drivers,
        "fixed_drivers": rep.fixed_drivers,
        "spare_drivers": rep.spare_drivers,
        "inferred_holidays": [d.isoformat() for d in rep.inferred_holidays],
        "group_rules": [
            {
                "group": g.group, "size": g.size,
                "rotation_step": g.rotation_step,
                "rotation_perm": g.rotation_perm,
                "rotation_support": round(g.rotation_support, 3),
                "reduction_mode": g.reduction_mode,
                "rest_slots": g.rest_slots,
                "display_mode": g.display_mode,
            }
            for g in rep.group_rules
        ],
        "recommendations": [
            {
                "key": r.key, "value": r.value,
                "confidence": round(r.confidence, 2),
                "evidence": r.evidence,
            }
            for r in rep.recommendations
        ],
    }


@app.post("/inspect")
async def inspect_endpoint(
    file: UploadFile = File(...),
    division: str = Form(""),
    sheets: str = Form(""),          # 비우면 마지막(가장 최근) 월 시트
    policy_json: str = Form(""),     # 비우면 저장된 회사 정책
    x_company_id: str = Header("default"),
):
    """검산 — 담당자가 이미 짜 놓은 배차표를 그대로 받아 규칙 위반을 찾는다.

    생성도 저장도 하지 않는다. 파일을 올리면 결과만 돌려주고 끝이다.
    이 무해함이 이 기능의 전부다 — 바꾸라고 하지 않으니 거절할 이유가 없다.
    """
    path = _save_upload(file)
    names = [s.strip() for s in sheets.split(",") if s.strip()] or _month_sheets(path)[-1:]
    if not names:
        raise HTTPException(400, "검사할 월 시트를 찾지 못했습니다")
    try:
        rosters = _load_rosters(path, division, names)
    except HTTPException:
        raise
    except Exception as ex:
        raise HTTPException(422, f"파싱 실패: {ex}")
    if not rosters:
        raise HTTPException(422, "시트에서 배차 데이터를 읽지 못했습니다")

    if policy_json.strip():
        pol = CompanyPolicy.from_dict(json.loads(policy_json))
    else:
        saved = _company_policy_path(x_company_id)
        pol = (
            CompanyPolicy.from_dict(json.loads(saved.read_text()))
            if saved.exists() else CompanyPolicy()
        )
    policy = pol.effective()
    reports = [inspect_roster(r, policy).to_dict() for r in rosters]
    return {"sheets": names, "reports": reports}


@app.post("/import")
async def import_endpoint(
    file: UploadFile = File(...),
    division: str = Form(""),
    sheets: str = Form(""),          # 비우면 마지막(가장 최근) 월 시트
):
    """그대로 가져오기 — 이미 짜 놓은 배차표를 solver 없이 읽어 cells 로 돌려준다.

    엔진으로 새로 짜기를 원치 않는 회사를 위한 경로다. 파일을 읽어 그대로
    돌려주면 백엔드가 /schedules/from-engine 로 저장하고, 그때부터 일일배차·
    차량별·기사별·게시 양식·기사앱·발행 안전검사가 전부 똑같이 동작한다.

    /generate 와 달리 과거 월(로테이션 이월)이 필요 없다 — 짜는 게 아니라
    읽는 것이므로 파일 한 장이면 된다.
    """
    path = _save_upload(file)
    names = [s.strip() for s in sheets.split(",") if s.strip()] or _month_sheets(path)[-1:]
    if not names:
        raise HTTPException(400, "가져올 월 시트를 찾지 못했습니다")
    try:
        rosters = _load_rosters(path, division, names)
    except HTTPException:
        raise
    except Exception as ex:
        raise HTTPException(422, f"파싱 실패: {ex}")
    if not rosters:
        raise HTTPException(422, "시트에서 배차 데이터를 읽지 못했습니다")

    roster = rosters[-1]
    # 시트는 찾았는데 한 칸도 못 읽은 경우 — 여기서 막지 않으면 빈 cells 가
    # 그대로 나가고, 백엔드는 "차량번호가 하나도 일치하지 않습니다"라며
    # 엉뚱하게 기초 데이터를 의심하게 만든다. 원인은 파일 양식이다.
    if not roster.entries:
        raise HTTPException(
            422,
            f"'{names[-1]}' 시트에서 배차 내용을 한 칸도 읽지 못했습니다. "
            "이 파일이 배차표 양식(행=차량, 칸=기사 이름)인지 확인해 주세요. "
            "Busync 에서 내보낸 파일이라면 '일별 상세' 시트가 함께 있어야 읽을 수 있습니다.",
        )
    cells: dict = defaultdict(dict)
    filled = 0
    for (d, v), e in sorted(roster.entries.items()):
        # 시트가 앞뒤 달을 걸쳐 있는 경우가 많다(게시용 주간표는 전월 말일부터
        # 시작). 그 달 것만 가져와야 5/31 이 6월 배차표에 감차로 섞이지 않는다.
        if (d.year, d.month) != (roster.year, roster.month):
            continue
        g = roster.group_of(v)
        am = e.am.driver or ""
        pm = e.pm.driver or ""
        # 감차 판정은 순번(slot_index) 유무로 하지 않는다 — 월간배차 양식은
        # 순번이 노선 고정값이라 휴차인 날도 값이 남아 있다. 그날 두 칸 모두
        # 기사가 없으면 안 나간 것으로 본다 (주간·월간 양식 모두에서 옳다).
        operating = bool(am or pm)
        if am:
            filled += 1
        if pm:
            filled += 1
        cells[d.isoformat()][v] = {
            "slot": e.slot_label,
            "display_slot": e.slot_index,
            # 그대로 가져온 표에는 언더라잉 로테이션 정보가 없다 — 표시 순번을 쓴다
            "am": am or e.am.raw,
            "pm": pm or e.pm.raw,
            "underlying": e.slot_index,
            "operating": operating,
            "group": g.name if g else None,
        }

    return {
        "year": roster.year,
        "month": roster.month,
        "sheets": names,
        "groups": [{"name": g.name, "vehicles": g.vehicles} for g in roster.groups],
        # 집계는 실제로 가져온(그 달) 범위 기준이어야 화면 숫자와 맞는다
        "vehicles": len({v for by in cells.values() for v in by}),
        "dates": len(cells),
        "filled_cells": filled,
        "drivers": sorted(roster.drivers()),
        "cells": cells,
    }


@app.post("/backtest")
async def backtest_endpoint(
    file: UploadFile = File(...),
    division: str = Form(""),
    prev_sheet: str = Form(...),
    target_sheet: str = Form(...),
    history_sheets: str = Form(""),
    holidays: str = Form(""),
    time_limit_s: float = Form(120.0),
    stage2: bool = Form(True),
):
    path = _save_upload(file)
    hols = {dt.date.fromisoformat(x) for x in holidays.split(",") if x.strip()}
    prev = parse_workbook_month(path, prev_sheet, division)
    target = parse_workbook_month(path, target_sheet, division)
    history = [
        parse_workbook_month(path, n.strip(), division)
        for n in history_sheets.split(",") if n.strip()
    ] + [prev]
    res1, _ = backtest_stage1(prev, target, hols)
    out = {
        "stage1": {
            "slot_match": res1.slot_match,
            "slot_total": res1.slot_total,
            "slot_rate": round(res1.slot_rate, 4),
            "mismatches": [
                [d.isoformat(), v, p, a]
                for d, v, p, a in res1.slot_mismatches[:50]
            ],
        }
    }
    if stage2:
        res2, asg = backtest_stage2(history, target, time_limit_s=time_limit_s)
        rep = run_audit(res2.problem, res2.assignment)
        out["stage2"] = {
            "cell_match": res2.cell_match,
            "cell_total": res2.cell_total,
            "cell_rate": round(res2.cell_rate, 4),
            "solver_status": asg.status,
            "audit_violations": len(rep.violations),
            "exceptions": [
                [d.isoformat(), v, s, got, want]
                for d, v, s, got, want in res2.cell_mismatches[:100]
            ],
        }
    return out


def _generation_response(result, year: int, month: int) -> dict:
    """생성 결과 → 응답 dict. 파일 경로와 JSON 경로가 같은 형식을 내도록 공유한다."""
    # 패턴(1단계)에서 언더라잉 슬롯·운행여부·그룹을 끌어와 셀에 붙인다.
    # 백엔드가 이 값들을 SchedulePattern으로 영속화해 다음 달 로테이션을 이어간다.
    pattern_meta: dict[tuple[dt.date, str], dict] = {}
    for gname, pat in result.patterns.items():
        for (d, v), cell in pat.items():
            pattern_meta[(d, v)] = {
                "underlying": cell.underlying_slot,
                "operating": cell.operating,
                "group": gname,
            }
    cells: dict = defaultdict(dict)
    for (d, v), e in sorted(result.roster.entries.items()):
        meta = pattern_meta.get((d, v), {})
        cells[d.isoformat()][v] = {
            "slot": e.slot_label,
            "display_slot": e.slot_index,
            "am": e.am.driver or e.am.raw,
            "pm": e.pm.driver or e.pm.raw,
            "underlying": meta.get("underlying"),
            "operating": meta.get("operating", True),
            "group": meta.get("group"),
        }
    return {
        "draft_id": _store_draft(result),
        "year": year,
        "month": month,
        "groups": [
            {"name": g.name, "vehicles": g.vehicles} for g in result.roster.groups
        ],
        "solver_status": result.assignment.status,
        "objective": result.assignment.objective,
        "warnings": result.warnings,
        "audit": {
            "ok": result.audit.ok,
            "violations": [
                {"rule": v.rule, "message": v.message} for v in result.audit.violations
            ],
        },
        "unfilled": [[d.isoformat(), v, s] for d, v, s in result.assignment.unfilled],
        "fairness": {
            "slot_balance_stdev": round(result.fairness.slot_balance_stdev, 2),
            "weekend_off_stdev": round(result.fairness.weekend_off_stdev, 2),
            "substitute_stdev": round(result.fairness.substitute_stdev, 2),
        },
        "cells": cells,
    }


@app.post("/generate-from-cells")
async def generate_from_cells_endpoint(
    payload: dict,
    x_company_id: str = Header("default"),
):
    """이미 저장된 지난달 배차표(cells)로 이번 달을 짠다 — 파일 업로드 없이.

    "지난달 배차표로 이번 달 짜기"에 엑셀 왕복이 필요할 이유가 없다. 지난달
    배차표는 Busync 안에 있고 순번(SchedulePattern)까지 저장돼 있다. 백엔드가
    그걸 cells 로 만들어 여기로 보내면 파일 경로와 똑같이 처리한다.

    payload: {year, month, history: [{year, month, cells, groups?}], division?,
              policy?, leaves?, home_config?, time_limit_s?}
    """
    year = int(payload.get("year") or 0)
    month = int(payload.get("month") or 0)
    if not (year and 1 <= month <= 12):
        raise HTTPException(400, "year/month 가 필요합니다")

    hist_in = payload.get("history") or []
    if not hist_in:
        raise HTTPException(400, "지난달 배차표(history)가 필요합니다")

    division = str(payload.get("division") or "")
    history = []
    for h in hist_in:
        r = roster_from_cells(
            h.get("cells") or {},
            int(h.get("year")),
            int(h.get("month")),
            division=division,
            groups=h.get("groups"),
        )
        if not r.entries:
            raise HTTPException(422, "지난달 배차표에서 배차 내용을 읽지 못했습니다")
        history.append(r)
    history.sort(key=lambda r: (r.year, r.month))

    if payload.get("policy"):
        policy = CompanyPolicy.from_dict(payload["policy"])
    else:
        saved = _company_policy_path(x_company_id)
        policy = (
            CompanyPolicy.from_dict(json.loads(saved.read_text()))
            if saved.exists() else CompanyPolicy()
        )
    leaves = {
        k: {dt.date.fromisoformat(x) for x in v}
        for k, v in (payload.get("leaves") or {}).items()
    }
    # 요일별 운행 대수는 **추론하지 말고 등록값을 쓴다**. 회사가 기초 데이터
    # (노선 > 평일/토/일·공휴일 대수)에 이미 정확히 적어 뒀는데, 지난달 실적에서
    # 되짚으면 평일 상시 감차를 공휴일로 오해해 대수가 뭉개진다.
    # {"16": {"weekday": 12, "sat": 11, "sunhol": 10, "fleet": 14}}
    operating_counts = payload.get("operating_counts") or None
    try:
        result = generate_month(
            history, policy, year, month,
            leaves=leaves,
            home_vehicle_config=payload.get("home_config") or None,
            time_limit_s=float(payload.get("time_limit_s") or 90.0),
            operating_counts=operating_counts,
            # mains_only: 기본 틀만 깔고 스페어 자리는 비워 둔다.
            # 담당자가 직접 채울지 엔진에 맡길지 고르게 하기 위함이다.
            mains_only=bool(payload.get("mains_only")),
        )
    except (ValueError, RuntimeError) as ex:
        raise HTTPException(422, str(ex))

    return _generation_response(result, year, month)


@app.post("/generate")
async def generate_endpoint(
    file: UploadFile = File(...),
    division: str = Form(""),
    year: int = Form(...),
    month: int = Form(...),
    history_sheets: str = Form(""),   # 비우면 최근 2개월 자동
    policy_json: str = Form(""),      # CompanyPolicy.to_dict() 형식. 비우면 저장된 회사 정책 사용
    leaves_json: str = Form("{}"),    # {기사명: ["YYYY-MM-DD", ...]}
    home_config_json: str = Form(""),  # {기사명: 차량} — 담당자 확정 월초 구성
    time_limit_s: float = Form(90.0),
    include_xlsx: bool = Form(False),
    x_company_id: str = Header("default"),
):
    path = _save_upload(file)
    names = [s.strip() for s in history_sheets.split(",") if s.strip()] or \
        _month_sheets(path)[-2:]
    if not names:
        raise HTTPException(400, "과거 월 시트를 찾지 못했습니다")
    try:
        history = _load_rosters(path, division, names)
    except HTTPException:
        raise
    except Exception as ex:
        raise HTTPException(422, f"파싱 실패: {ex}")
    if policy_json.strip():
        policy = CompanyPolicy.from_dict(json.loads(policy_json))
    else:
        saved = _company_policy_path(x_company_id)
        policy = (
            CompanyPolicy.from_dict(json.loads(saved.read_text()))
            if saved.exists() else CompanyPolicy()
        )
    leaves = {
        k: {dt.date.fromisoformat(x) for x in v}
        for k, v in json.loads(leaves_json).items()
    }
    home_cfg = json.loads(home_config_json) if home_config_json.strip() else None
    try:
        result = generate_month(
            history, policy, year, month,
            leaves=leaves, home_vehicle_config=home_cfg,
            time_limit_s=time_limit_s,
        )
    except (ValueError, RuntimeError) as ex:
        raise HTTPException(422, str(ex))

    out = _generation_response(result, year, month)
    if include_xlsx:
        from busync_engine.renderer import render_weekly_xlsx

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            render_weekly_xlsx(result.roster, f.name)
            f.seek(0)
            out["xlsx_base64"] = base64.b64encode(open(f.name, "rb").read()).decode()
    return out


@app.get("/draft/{draft_id}/explain")
def explain_endpoint(draft_id: str, date: str, vehicle: str, shift: str):
    """셀 배정 근거 — "왜 이 기사가 이 슬롯인가" (스펙 0)."""
    from busync_engine.explain import explain_cell

    draft = _get_draft(draft_id)
    result = draft["result"]
    exp = explain_cell(
        result.problem, result.assignment,
        dt.date.fromisoformat(date), vehicle, shift.upper(),
    )
    return {
        "driver": exp.driver,
        "summary": exp.summary,
        "reasons": [
            {"code": r.code, "text": r.text, "weight": r.weight}
            for r in exp.reasons
        ],
        "alternatives": [
            {"driver": k, "extra_penalty": c} for k, c in exp.alternatives
        ],
    }


@app.get("/draft/{draft_id}/explain-driver")
def explain_driver_endpoint(draft_id: str, driver: str):
    """기사 월간 배정 설명 — 기사 앱 '내 배차 설명' 화면."""
    from busync_engine.explain import explain_driver_month

    draft = _get_draft(draft_id)
    result = draft["result"]
    return {
        "driver": driver,
        "lines": explain_driver_month(result.problem, result.assignment, driver),
    }


@app.post("/draft/{draft_id}/absence")
def absence_endpoint(
    draft_id: str,
    date: str = Form(...),
    vehicle: str = Form(...),
    shift: str = Form(...),
):
    """당일 결원 신고 → 국소 수리 후보 추천 (전체 재생성 금지 — 스펙 4-3)."""
    from busync_engine.repair import mark_absent, suggest_repair

    draft = _get_draft(draft_id)
    result = draft["result"]
    d = dt.date.fromisoformat(date)
    absent = mark_absent(result.assignment, d, vehicle, shift.upper())
    sug = suggest_repair(
        result.problem, result.assignment, d, vehicle, shift.upper()
    )
    return {
        "absent_driver": absent,
        "top_candidates": [
            {"driver": c.driver, "score": round(c.score, 1),
             "reasons": c.reasons}
            for c in sug.top
        ],
        "blocked": [
            {"driver": c.driver, "blocking": c.blocking}
            for c in sug.candidates if not c.eligible
        ][:10],
    }


@app.post("/draft/{draft_id}/repair")
def repair_endpoint(
    draft_id: str,
    date: str = Form(...),
    vehicle: str = Form(...),
    shift: str = Form(...),
    driver: str = Form(...),
    reason: str = Form(""),
):
    """대체 확정 — 해당 셀만 교체, 이력은 다음 달 공정성 계산에 반영."""
    from busync_engine.repair import apply_repair

    draft = _get_draft(draft_id)
    result = draft["result"]
    rec = apply_repair(
        result.assignment, dt.date.fromisoformat(date), vehicle,
        shift.upper(), driver, reason=reason,
    )
    draft["repairs"].append(rec)
    return {
        "changed": {
            "date": rec.date.isoformat(), "vehicle": rec.vehicle,
            "shift": rec.shift, "removed": rec.removed, "added": rec.added,
        },
        "repair_count": len(draft["repairs"]),
    }


@app.post("/leave/triage")
def leave_triage_endpoint(payload: dict):
    """휴무신청 분류 — 정원 내 자동 승낙, 초과분은 호혜성 스코어로 검토 대기.

    payload 예:
    {
      "daily_cap": 6,
      "requests": [{"driver": "김학순", "date": "2026-08-15",
                    "type": "휴", "requested_at": "2026-08-01T09:00:00"}],
      "counters": {"김학순": {"requested": 1, "accepted": 6, "rejected": 0}},
      "available_by_date": {"2026-08-15": 3},
      "approval_rates": {"김학순": 0.9}
    }
    """
    from busync_engine.leave import (
        LeaveRequest, SubstituteCounter, triage,
    )
    from busync_engine.models import LeaveType

    reqs = []
    for r in payload.get("requests", []):
        try:
            lt = LeaveType(r.get("type", "휴"))
        except ValueError:
            lt = LeaveType.REGULAR
        reqs.append(LeaveRequest(
            driver=r["driver"],
            date=dt.date.fromisoformat(r["date"]),
            leave_type=lt,
            requested_at=(
                dt.datetime.fromisoformat(r["requested_at"])
                if r.get("requested_at") else None
            ),
            reason=r.get("reason", ""),
        ))
    counters = {
        k: SubstituteCounter(**v)
        for k, v in payload.get("counters", {}).items()
    }
    available = {
        dt.date.fromisoformat(k): v
        for k, v in payload.get("available_by_date", {}).items()
    }
    decisions = triage(
        reqs, counters,
        daily_cap=int(payload.get("daily_cap", 6)),
        available_by_date=available,
        recent_approval_rate=payload.get("approval_rates"),
    )
    return {
        "days": [
            {
                "date": dec.date.isoformat(),
                "capacity": dec.capacity,
                "auto_approved": [
                    {"driver": r.driver, "note": r.decision_note}
                    for r in dec.auto_approved
                ],
                "needs_review": [
                    {
                        "driver": s.request.driver,
                        "score": round(s.score, 2),
                        "evidence": s.evidence,
                        "breakdown": s.breakdown,
                    }
                    for s in dec.needs_review
                ],
            }
            for dec in decisions
        ]
    }


@app.get("/leave/annual")
def annual_leave_endpoint(hire_date: str, as_of: str = ""):
    """근로기준법 연차 자동계산 (스펙 2.8)."""
    from busync_engine.leave import annual_leave_days

    ref = dt.date.fromisoformat(as_of) if as_of else dt.date.today()
    hd = dt.date.fromisoformat(hire_date)
    return {
        "hire_date": hire_date,
        "as_of": ref.isoformat(),
        "annual_leave_days": annual_leave_days(hd, ref),
    }


@app.get("/draft/{draft_id}/xlsx")
def draft_xlsx_endpoint(draft_id: str):
    """초안을 게시 양식 xlsx로 다운로드 — 수리(대체) 반영 상태로 재렌더링."""
    from fastapi.responses import Response

    from busync_engine.generate import _to_roster
    from busync_engine.renderer import render_weekly_xlsx
    from busync_engine.rotation import DisplayMode

    draft = _get_draft(draft_id)
    result = draft["result"]
    base = result.roster
    display_of = {g.name: DisplayMode.KEEP for g in base.groups}
    roster = _to_roster(
        result.assignment, result.patterns, base.groups, display_of,
        base.year, base.month, base.division,
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        render_weekly_xlsx(roster, f.name)
        f.seek(0)
        content = open(f.name, "rb").read()
    filename = f"dispatch-{base.year}-{base.month:02d}.xlsx"
    return Response(
        content=content,
        media_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )
